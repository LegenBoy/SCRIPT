import streamlit as st
import pandas as pd
import os
import re
import shutil
import copy
import warnings
import tempfile
from io import BytesIO
from openpyxl import load_workbook

# Ignorar avisos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ================= CONFIGURAÇÕES DE LAYOUT =================
NOME_MODELO_PADRAO = "MODELO.xlsx" 
NOME_BASE_CEPS = "MODELO(2).xlsx"

LINHA_INICIAL_DADOS = 7  
PASSO_ENTRE_ROTAS = 4    
MARGEM_COLUNAS = 2       

# ================= FUNÇÕES DE ESTILO =================
def formatar_cidade(texto):
    if pd.isna(texto) or str(texto).strip() == "":
        return None
    texto = str(texto)
    padrao = r"^(\d+)-(.*?)[/]"
    match = re.search(padrao, texto)
    if match:
        return f"{match.group(2).strip()} - {match.group(1)}"
    return texto

def copiar_estilo(celula_origem, celula_destino):
    if celula_origem.has_style:
        celula_destino.font = copy.copy(celula_origem.font)
        celula_destino.border = copy.copy(celula_origem.border)
        celula_destino.fill = copy.copy(celula_origem.fill)
        celula_destino.number_format = copy.copy(celula_origem.number_format)
        celula_destino.protection = copy.copy(celula_origem.protection)
        celula_destino.alignment = copy.copy(celula_origem.alignment)

def replicar_bloco_formatacao(ws, linha_base, linha_nova_inicio, altura_bloco):
    max_col = ws.max_column
    for i in range(altura_bloco):
        src_row = linha_base + i
        dst_row = linha_nova_inicio + i
        for col in range(1, max_col + 1):
            src_cell = ws.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)
            copiar_estilo(src_cell, dst_cell)

    merges_origem = [
        rng for rng in ws.merged_cells.ranges
        if rng.min_row >= linha_base and rng.max_row < (linha_base + altura_bloco)
    ]
    for rng in merges_origem:
        offset = linha_nova_inicio - linha_base
        ws.merge_cells(
            start_row=rng.min_row + offset, start_column=rng.min_col,
            end_row=rng.max_row + offset, end_column=rng.max_col
        )

def preparar_estrutura_linhas(ws, total_rotas):
    ultima_linha_necessaria = LINHA_INICIAL_DADOS + (total_rotas * PASSO_ENTRE_ROTAS)
    max_row_excel = ws.max_row
    
    # Lógica de verificação: olha se existem linhas formatadas suficientes
    linha_atual_verificacao = LINHA_INICIAL_DADOS
    rotas_formatadas_existentes = 0
    
    # Percorre para contar quantas rotas já têm "caixinha azul" (borda esquerda)
    while True:
        if linha_atual_verificacao > max_row_excel + 100: # Limite de segurança para loop infinito
            break
            
        cell = ws.cell(row=linha_atual_verificacao, column=2)
        # Se não tiver borda ou estilo, assumimos que acabou o template
        if not cell.border.left.style: 
            break
            
        rotas_formatadas_existentes += 1
        linha_atual_verificacao += PASSO_ENTRE_ROTAS
    
    # AQUI ESTAVA O BUG: Aumentamos a margem de segurança de +5 para +50
    # Isso garante que ele crie muitas linhas a mais antes de escrever
    if rotas_formatadas_existentes < total_rotas + 10:
        rotas_faltantes = (total_rotas - rotas_formatadas_existentes) + 50 
        linha_destino = linha_atual_verificacao
        
        for _ in range(rotas_faltantes):
            replicar_bloco_formatacao(ws, LINHA_INICIAL_DADOS, linha_destino, PASSO_ENTRE_ROTAS)
            linha_destino += PASSO_ENTRE_ROTAS

def obter_range_mesclado(ws, row, col):
    for merged_range in ws.merged_cells.ranges:
        if (row >= merged_range.min_row and row <= merged_range.max_row and
            col >= merged_range.min_col and col <= merged_range.max_col):
            return merged_range
    return None

def escrever_valor(ws, row, col_inicial, valor):
    col_atual = col_inicial
    while True:
        merged_range = obter_range_mesclado(ws, row, col_atual)
        if merged_range:
            if row == merged_range.min_row and col_atual == merged_range.min_col:
                celula = ws.cell(row, col_atual)
                celula.value = valor
                return merged_range.max_col + 1
            else:
                col_atual = merged_range.max_col + 1
                continue
        else:
            celula = ws.cell(row, col_atual)
            celula.value = valor
            return col_atual + 1

def limpar_sobras_total(ws, ultima_linha_usada, ultima_coluna_usada):
    linha_inicio_corte = ultima_linha_usada + PASSO_ENTRE_ROTAS
    max_row = ws.max_row
    # Margem extra antes de cortar para evitar cortar a última borda
    if max_row >= linha_inicio_corte:
        qtd_linhas_apagar = (max_row - linha_inicio_corte) + 200
        ws.delete_rows(linha_inicio_corte, qtd_linhas_apagar)

    coluna_inicio_corte = ultima_coluna_usada + MARGEM_COLUNAS + 1
    max_col = ws.max_column
    if max_col >= coluna_inicio_corte:
        qtd_cols_apagar = (max_col - coluna_inicio_corte) + 50
        ws.delete_cols(coluna_inicio_corte, qtd_cols_apagar)

def ajustar_largura_colunas(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if max_length > 0:
            ws.column_dimensions[column].width = max_length + 3

def processar_arquivo(caminho_arquivo, dataframe):
    try:
        wb = load_workbook(caminho_arquivo)
        if "MODELO" in wb.sheetnames: ws = wb["MODELO"]
        elif "IMPRESSÃO" in wb.sheetnames: ws = wb["IMPRESSÃO"]
        else: ws = wb.active
    except Exception as e:
        st.error(f"Erro ao abrir modelo: {e}")
        return False

    preparar_estrutura_linhas(ws, len(dataframe))

    linha_atual = LINHA_INICIAL_DADOS
    max_col_global = 2 

    for index, row in dataframe.iterrows():
        coluna_cursor = 2
        for i in range(1, 13):
            col_nome = f"filial{i}/cubagem"
            if col_nome in dataframe.columns:
                valor = row[col_nome]
                cidade_fmt = formatar_cidade(valor)
                if cidade_fmt:
                    coluna_cursor = escrever_valor(ws, linha_atual, coluna_cursor, cidade_fmt)
        
        coluna_usada_nesta_linha = coluna_cursor - 1
        if coluna_usada_nesta_linha > max_col_global:
            max_col_global = coluna_usada_nesta_linha
            
        linha_atual += PASSO_ENTRE_ROTAS

    ultima_linha_real = linha_atual - PASSO_ENTRE_ROTAS
    limpar_sobras_total(ws, ultima_linha_real, max_col_global)
    ajustar_largura_colunas(ws)
    
    wb.save(caminho_arquivo)
    return True

# ================= INTERFACE STREAMLIT =================

def main():
    st.set_page_config(page_title="Gerador de Rotas", layout="centered")
    st.title("🚛 Previsão de Descarga")

    # Verifica se o MODELO.xlsx existe
    if not os.path.exists(NOME_MODELO_PADRAO):
        st.error(f"ERRO CRÍTICO: '{NOME_MODELO_PADRAO}' não encontrado.")
        return

    # --- INICIALIZAÇÃO DO ESTADO (SESSION STATE) ---
    # Isso garante que os dados fiquem salvos mesmo após clicar em baixar
    if 'arquivos_prontos' not in st.session_state:
        st.session_state['arquivos_prontos'] = []

    # Upload
    arquivo_upload = st.file_uploader("Selecione a planilha de dados (Excel)", type=["xlsx"])

    # Botão de resetar (caso queira limpar a tela e começar de novo)
    if st.session_state['arquivos_prontos']:
        if st.button("🔄 Limpar e Processar Nova Planilha"):
            st.session_state['arquivos_prontos'] = []
            st.rerun()

    if arquivo_upload is not None:
        # Só mostra o botão Processar se ainda não tiver processado
        if not st.session_state['arquivos_prontos']:
            if st.button("Processar Arquivos"):
                with st.spinner('Processando... Aguarde.'):
                    with tempfile.TemporaryDirectory() as tmpdirname:
                        
                        # Salva input
                        caminho_input = os.path.join(tmpdirname, "input.xlsx")
                        with open(caminho_input, "wb") as f:
                            f.write(arquivo_upload.getbuffer())

                        try:
                            df = pd.read_excel(caminho_input)
                            df.columns = [str(c).strip().lower() for c in df.columns]
                        except Exception as e:
                            st.error(f"Erro ao ler arquivo: {e}")
                            return

                        col_transp = next((c for c in df.columns if "transportadora" in c), None)
                        if not col_transp:
                            st.error("Erro: Coluna 'transportadora' não encontrada.")
                            return

                        # --- VERIFICAÇÃO DE FRETE RETORNO ---
                        if os.path.exists(NOME_BASE_CEPS):
                            try:
                                df_base = pd.read_excel(NOME_BASE_CEPS)
                                # Verifica se tem colunas suficientes (D=3, E=4)
                                if df_base.shape[1] >= 5:
                                    def clean_cep(c):
                                        return re.sub(r'\D', '', str(c)) if pd.notna(c) else ''

                                    base_map = {}
                                    for _, row_b in df_base.iterrows():
                                        forn = row_b.iloc[3] # Coluna D (índice 3)
                                        cep_b = row_b.iloc[4] # Coluna E (índice 4)
                                        cep_clean = clean_cep(cep_b)
                                        if cep_clean:
                                            if cep_clean not in base_map:
                                                base_map[cep_clean] = set()
                                            base_map[cep_clean].add(str(forn))
                                    
                                    # Verifica input (Coluna AA = 27ª coluna = índice 26)
                                    if df.shape[1] > 26:
                                        ceps_alertas = []
                                        col_aa = df.iloc[:, 26]
                                        for idx_r, val_aa in col_aa.items():
                                            c_in = clean_cep(val_aa)
                                            if c_in and c_in in base_map:
                                                fornecedores_encontrados = ", ".join(base_map[c_in])
                                                ceps_alertas.append(f"CEP {val_aa} (Linha {idx_r+2}) -> Fornecedor: {fornecedores_encontrados}")
                                        
                                        if ceps_alertas:
                                            st.warning("🚨 **ATENÇÃO: FRETE RETORNO IDENTIFICADO**")
                                            st.write("Os seguintes CEPs da coluna AA constam na base de Frete Retorno. **Verifique qual o FORNECEDOR / CD CORRETO:**")
                                            st.dataframe(pd.DataFrame(ceps_alertas, columns=["Detalhes do Alerta"]), hide_index=True)
                            except Exception as e:
                                st.error(f"Erro na verificação de frete retorno: {e}")

                        lista_temp_arquivos = []

                        # --- GERA O GERAL ---
                        caminho_geral = os.path.join(tmpdirname, "GERAL_ROTAS.xlsx")
                        shutil.copy(NOME_MODELO_PADRAO, caminho_geral)
                        
                        if processar_arquivo(caminho_geral, df):
                            # Lê o arquivo gerado para a memória (BytesIO)
                            with open(caminho_geral, "rb") as f:
                                dados_binarios = f.read()
                            lista_temp_arquivos.append({
                                "nome": "GERAL_ROTAS.xlsx",
                                "dados": dados_binarios
                            })

                        # --- GERA POR TRANSPORTADORA ---
                        for transp, dados in df.groupby(col_transp):
                            if pd.isna(transp): continue
                            nome_arquivo = str(transp).replace("/", "-").replace("\\", "").strip() + ".xlsx"
                            caminho_transp = os.path.join(tmpdirname, nome_arquivo)
                            
                            shutil.copy(NOME_MODELO_PADRAO, caminho_transp)
                            if processar_arquivo(caminho_transp, dados):
                                with open(caminho_transp, "rb") as f:
                                    dados_binarios = f.read()
                                lista_temp_arquivos.append({
                                    "nome": nome_arquivo,
                                    "dados": dados_binarios
                                })

                        # Salva tudo na sessão
                        st.session_state['arquivos_prontos'] = lista_temp_arquivos
                        st.success("Concluído!")
                        st.rerun() # Recarrega a página para mostrar os botões

    # --- ÁREA DE DOWNLOAD (FORA DO IF DO BOTÃO) ---
    # Isso garante que os botões persistam na tela
    if st.session_state['arquivos_prontos']:
        st.divider()
        st.subheader("📂 Arquivos Gerados:")
        
        for item in st.session_state['arquivos_prontos']:
            st.download_button(
                label=f"📥 Baixar {item['nome']}",
                data=item['dados'],
                file_name=item['nome'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=item['nome'] # Chave única para não dar conflito
            )

if __name__ == "__main__":
    main()
